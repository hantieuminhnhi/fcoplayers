import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import random
from datetime import datetime

# Tạo dữ liệu mẫu cho đội hình FCOline
teams = [
    "FC Barcelona", "Real Madrid", "Manchester City", "Liverpool", 
    "Bayern Munich", "PSG", "Manchester United", "Arsenal"
]

positions = ["ST", "LW", "RW", "CAM", "CM", "CDM", "LB", "RB", "CB", "GK"]
player_names = {
    "ST": ["Messi", "Ronaldo", "Haaland", "Mbappe", "Lewandowski", "Kane"],
    "LW": ["Neymar", "Vinicius Jr", "Rashford", "Son", "Mane"],
    "RW": ["Salah", "Messi", "Dembele", "Sancho", "Di Maria"],
    "CAM": ["De Bruyne", "Bruno Fernandes", "Muller", "Odegaard", "Griezmann"],
    "CM": ["Modric", "Kroos", "Kimmich", "Pedri", "Bellingham"],
    "CDM": ["Casemiro", "Rodri", "Rice", "Fabinho", "Busquets"],
    "LB": ["Robertson", "Davies", "Mendy", "Cancelo", "Theo Hernandez"],
    "RB": ["Alexander-Arnold", "Walker", "Hakimi", "James", "Carvajal"],
    "CB": ["Van Dijk", "Rudiger", "Marquinhos", "Dias", "Salah"],
    "GK": ["Courtois", "Alisson", "Neuer", "Ter Stegen", "Oblak"]
}

# Tạo danh sách cầu thủ
players = []
for i in range(1, 24):  # Tạo 23 cầu thủ (11 chính + 12 dự bị)
    team = random.choice(teams)
    position = random.choice(positions)
    name = random.choice(player_names[position])
    # Đảm bảo tên không trùng quá nhiều
    if any(p['Name'] == name and p['Position'] == position for p in players):
        name = f"{name} {i}"
    
    players.append({
        'ID': i,
        'Name': name,
        'Position': position,
        'Overall': random.randint(75, 105),
        'Team': team,
        'Form': random.choice(['↑', '→', '↓', '↑↑', '↓↓']),
        'Value (M)': round(random.uniform(5, 150), 1),
        'Contract': random.randint(30, 365),
        'Goals': random.randint(0, 25),
        'Assists': random.randint(0, 20),
        'Matches': random.randint(5, 50)
    })

# Chuyển thành DataFrame
df_players = pd.DataFrame(players)

# Tạo dữ liệu chiến thuật
tactics = {
    'Formation': ['4-2-3-1', '4-3-3', '4-4-2', '3-5-2', '5-3-2', '4-1-2-1-2'],
    'Defense Style': ['Balanced', 'Pressure', 'Drop Back', 'Heavy Press', 'Counter'],
    'Attack Style': ['Possession', 'Direct', 'Wide', 'Central', 'Fast Build'],
    'Defense Width': [4, 5, 3, 4, 5, 4],
    'Attack Width': [5, 6, 4, 7, 5, 6],
    'Players In Box': [4, 5, 3, 4, 4, 5]
}

df_tactics = pd.DataFrame(tactics)

# Tạo dữ liệu lịch sử trận đấu
matches = []
for i in range(1, 21):
    matches.append({
        'Match ID': i,
        'Opponent': random.choice(teams),
        'Score': f"{random.randint(0, 5)}-{random.randint(0, 5)}",
        'Result': random.choice(['Win', 'Loss', 'Draw']),
        'Competition': random.choice(['VSA', 'Manager Mode', 'Friendly', 'Tournament']),
        'Date': f"2024-{random.randint(1, 12):02d}-{random.randint(1, 28):02d}"
    })

df_matches = pd.DataFrame(matches)

# Tạo file Excel với nhiều sheet
excel_file = "FCOline_Team_Manager.xlsx"

with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    df_players.to_excel(writer, sheet_name='Squad', index=False)
    df_tactics.to_excel(writer, sheet_name='Tactics', index=False)
    df_matches.to_excel(writer, sheet_name='Match History', index=False)
    
    # Tạo sheet Transfer Market
    transfer_data = []
    for team in teams:
        transfer_data.append({
            'Player': f"Player from {team}",
            'Position': random.choice(positions),
            'Overall': random.randint(70, 100),
            'Price (M)': round(random.uniform(1, 80), 1),
            'Team': team,
            'Status': random.choice(['Available', 'Bidding', 'Sold'])
        })
    df_transfer = pd.DataFrame(transfer_data)
    df_transfer.to_excel(writer, sheet_name='Transfer Market', index=False)
    
    # Tạo sheet Stats
    stats_data = []
    positions_stats = ['ST', 'MF', 'DF', 'GK']
    for pos in positions_stats:
        stats_data.append({
            'Position': pos,
            'Total Goals': random.randint(50, 200),
            'Total Assists': random.randint(40, 150),
            'Avg Rating': round(random.uniform(6.5, 8.5), 1),
            'MVP Count': random.randint(5, 20)
        })
    df_stats = pd.DataFrame(stats_data)
    df_stats.to_excel(writer, sheet_name='Team Stats', index=False)

# Định dạng file Excel
wb = load_workbook(excel_file)

# Định dạng cho từng sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Định dạng header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Tự động điều chỉnh độ rộng cột
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

# Định dạng sheet Squad đặc biệt
ws_squad = wb['Squad']
# Thêm màu sắc cho Form
for row in ws_squad.iter_rows(min_row=2, max_row=ws_squad.max_row, min_col=6, max_col=6):
    for cell in row:
        if cell.value == '↑↑':
            cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        elif cell.value == '↑':
            cell.fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
        elif cell.value == '→':
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        elif cell.value in ['↓', '↓↓']:
            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

# Thêm bảng tổng hợp vào sheet Squad
ws_squad['K1'] = "TEAM STATISTICS"
ws_squad['K1'].font = Font(bold=True, size=14)
ws_squad['K2'] = "Total Players:"
ws_squad['L2'] = len(players)
ws_squad['K3'] = "Average Rating:"
avg_rating = sum(p['Overall'] for p in players) / len(players)
ws_squad['L3'] = round(avg_rating, 1)
ws_squad['K4'] = "Top Scorer:"
top_scorer = max(players, key=lambda x: x['Goals'])
ws_squad['L4'] = f"{top_scorer['Name']} ({top_scorer['Goals']} goals)"
ws_squad['K5'] = "Most Assists:"
top_assist = max(players, key=lambda x: x['Assists'])
ws_squad['L5'] = f"{top_assist['Name']} ({top_assist['Assists']} assists)"

# Định dạng bảng tổng hợp
for row in range(2, 6):
    ws_squad[f'K{row}'].font = Font(bold=True)
    ws_squad[f'L{row}'].alignment = Alignment(horizontal="left")

wb.save(excel_file)

print(f"✅ Đã tạo file Excel thành công: {excel_file}")
print("\n📋 Các sheet trong file:")
print("1. Squad - Danh sách đội hình với chỉ số và phong độ")
print("2. Tactics - Các chiến thuật có thể áp dụng")
print("3. Match History - Lịch sử các trận đấu")
print("4. Transfer Market - Thị trường chuyển nhượng")
print("5. Team Stats - Thống kê đội bóng")
print("\n🎮 File này giúp bạn quản lý đội hình FCOline hiệu quả!")